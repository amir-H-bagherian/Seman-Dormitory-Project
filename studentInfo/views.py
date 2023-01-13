from django.shortcuts import render, redirect
from .models import User
import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login


def read_from_excel():
    
    users = {}
    dataframe = openpyxl.load_workbook("Book1.xlsx")
    sheet = dataframe.active
    
    for i in range(1, sheet.max_row + 1):
        lst = []
        for j in range(1, sheet.max_column + 1):
            cell_obj = sheet.cell(row=i, column=j)
            lst.append(cell_obj.value)
        users.update({lst[0]: lst[1]})
        
    return users


def login_user(request):
    
    if request.method == 'POST':
        student_id = int(request.POST.get('student_id'))
        national_code = int(request.POST.get('national_code'))
        email = request.POST.get('email')
        users = read_from_excel()
        
        # If student has registered before and wants to regsiter again for whatever reason
        # Delete the student's record and rebuild a record for the student.
        # We do this since we do not have and do not want to have acounts for our users.
        try:
            user = User.objects.get(national_code=national_code)
            user.delete()
        except:
            pass
        
        if national_code in users.keys() and users[national_code] == student_id:
            
            new_user = User(national_code=national_code, student_id=student_id, email=email)
            new_user.save()
            login(request, new_user)

            return redirect('homepage') # go to the next page
        else:
            messages.error(request, "شماره دانشجویی یا کدملی وارد شده نامعتبر است")
    
    return render(request, 'index.html')


@login_required(login_url='index')
def home_page(request):
    return render(request, 'HomePage.html')


def farzanegan_page(request):
    return render(request, 'farzanegan.html')


def farhikhtegan_page(request):
    return render(request, 'farhikhtegan.html')


def kosar_page(request):
    return render(request, 'kosar.html')


@login_required(login_url='index')
def disabled_request_page(request):
    return render(request, 'disabled.html')


@login_required(login_url='index')
def shared_room(request):
    return render(request, 'request-shared-room.html')


@login_required(login_url='index')
def pre_exam_page(request):
    return render(request, 'pre-exam.html')


@login_required(login_url='index')
def submit_user_type(request):
    personal_type = request.POST.get('personalType')
    user = request.user
    user.personal_type = personal_type
    user.save()
    return redirect('behavior-test')


@login_required(login_url='index')
def behavior_test_page(request):
    context = {}
    return render(request, 'behavior-test.html', context)


@login_required(login_url='index')  
def process_user_lifestyle(request):
    user_bed_time_option = request.POST.get('bedTime')
    user_cigarette_option = request.POST.get('cigaretteStatus')
    user_tidyness_option = request.POST.get('tidynessStatus')
    
    bed_time_options = {
        'opt1': 'before12',
        'opt2': 'before2',
        'opt3': 'no-exact-time'}
    cigarette_options = {
        'opt1': True,
        'opt2': True,
        'opt3': False
    }
    tidyness_options = {
        'opt1': 'tidy',
        'opt2': 'semi-tidy',
        'opt3': 'untidy'
    }
    
    request.user.bed_time = bed_time_options[user_bed_time_option]
    request.user.cigarette_status = cigarette_options[user_cigarette_option]
    request.user.tidyness_status = tidyness_options[user_tidyness_option]
    request.user.save()
    return redirect('final-page')